#!/usr/bin/env python3
"""
Carton Sticker Generator — core engine.

The drawing for each sheet type lives in its OWN file so each can be edited
independently:
    layout_eight.py — 8-copy sheet  (sticker 4"   x 2.5")
    layout_four.py  — 4-copy sheet  (sticker 4.1" x 5")
    layout_two.py   — 2-copy sheet  (sticker 5"   x 5")
    layout_one.py   — 1-copy sheet  (sticker 6"   x 6")
Shared settings (font, oval, right-shift, letter in the oval) are in common.py.

Usage:
    python sticker_generator.py <po_file.xlsx> [more.xlsx ...]
"""

import sys, re, os, math
import zipfile
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

import layout_eight, layout_four, layout_two, layout_one
import common
from common import FONT, TEXT_HSCALE, CIRCLE_LETTER  # re-export for app use

# ----------------------------------------------------------------------
# Which layout for which carton Height (cm). Checked top-down.
COPY_RULES = [
    (18, 8),
    (26, 4),
    (30, 2),
    (9999, 1),
]

LAYOUTS = {8: layout_eight, 4: layout_four, 2: layout_two, 1: layout_one}

PAGE_SIZE = A4

# Each carton needs this many identical stickers (2 = one per side, etc.)
STICKERS_PER_CARTON = 2

def copies_for_height(h):
    for max_h, copies in COPY_RULES:
        if h <= max_h:
            return copies
    return 1

MEAS_RE = re.compile(r"L\s*(\d+(?:\.\d+)?)\s*X\s*W\s*(\d+(?:\.\d+)?)\s*X\s*H\s*(\d+(?:\.\d+)?)", re.I)
ITEM_RE = re.compile(r"^\d{3}-\d{6}\(\d+-\d+\)$")          # 341-489933(71-06)
VOL_RE  = re.compile(r"^\(([\d.]+)\s*M3\)$", re.I)          # (0.059M3)
PCS_RE  = re.compile(r"^(\d+)\s*PCS$", re.I)                # 16PCS
QTY_RE  = re.compile(r"Qty\s*Required\s*is\s*:?\s*(\d+)\s*box", re.I)
SIZE_NAMES = {"XS","S","M","L","XL","XXL","2XL","3XL","4XL","5XL","6XL","XXXL"}

def iter_cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and str(c.value).strip() != "":
                yield c

def cell_map(ws):
    """{(row,col): stripped string value}"""
    return {(c.row, c.column): str(c.value).strip() for c in iter_cells(ws)}

def find_first(cm, pred):
    for (r, col), v in sorted(cm.items()):
        if pred(v):
            return (r, col), v
    return None, None

def parse_measurement(cm):
    for _, v in cm.items():
        m = MEAS_RE.search(v)
        if m:
            return float(m.group(1)), float(m.group(2)), float(m.group(3))
    return None, None, None

def parse_common(cm):
    d = {}
    _, d["item_no"] = find_first(cm, lambda v: ITEM_RE.match(v))
    pos, pcs = find_first(cm, lambda v: PCS_RE.match(v))
    d["pcs"] = int(PCS_RE.match(pcs).group(1)) if pcs else None
    _, vol = find_first(cm, lambda v: VOL_RE.match(v))
    d["volume"] = vol
    d["L"], d["W"], d["H"] = parse_measurement(cm)
    _, po = find_first(cm, lambda v: v.upper().startswith(":P0"))
    d["po_no"] = po.lstrip(":") if po else None
    _, dest = find_first(cm, lambda v: re.match(r"^\d{3,5}/[A-Z]{2,3}$", v))
    d["dest"] = dest
    qty = None
    for _, v in cm.items():
        m = QTY_RE.search(v)
        if m:
            qty = int(m.group(1))
    d["qty_boxes"] = qty
    # description: longest text line that isn't a label/known field
    labels = ("width side","length side","margin","rfid","po no","gross","net",
              "measurement","country","qty required","carton qty","colors","pattern","colour")
    best = ""
    for _, v in cm.items():
        lv = v.lower()
        if any(k in lv for k in labels):
            continue
        if ITEM_RE.match(v) or PCS_RE.match(v) or VOL_RE.match(v) or v.startswith(":"):
            continue
        if len(v) > len(best) and re.search(r"[a-zA-Z]{4,}", v) and not re.match(r"^\d", v):
            best = v
    d["description"] = best or None
    return d

def parse_ratio_sheet(ws):
    """One sheet = one assortment carton design."""
    cm = cell_map(ws)
    d = parse_common(cm)
    d["type"] = "ratio"
    # assort code: cell directly below item_no (e.g. I017 / T008 / X053 / E031)
    pos, _ = find_first(cm, lambda v: ITEM_RE.match(v))
    d["code"] = None
    if pos:
        r, c = pos
        for dr in (1, 2):
            v = cm.get((r + dr, c))
            if v and re.match(r"^[A-Z]\d{3}$", v):
                d["code"] = v
                break
    if not d["code"]:
        d["code"] = ws.title.split()[0]
    # size table: rows where col X = size name, col X+1.. = numbers; "Total" row too
    table_rows = []
    first_size_pos = None
    for (r, c), v in sorted(cm.items()):
        if v.upper() in SIZE_NAMES or v.lower() == "total":
            nums = []
            cc = c + 1
            while (r, cc) in cm and re.match(r"^\d+$", cm[(r, cc)]):
                nums.append(int(cm[(r, cc)]))
                cc += 1
            if nums:
                if first_size_pos is None:
                    first_size_pos = (r, c)
                table_rows.append((v if v.lower() != "total" else "TOTAL", nums))
    # color header codes (e.g. 41, 50) sitting 1-4 rows above the first size row
    colors = []
    if first_size_pos:
        r0, c0 = first_size_pos
        for dr in range(1, 5):
            row_codes = []
            cc = c0 + 1
            while (r0 - dr, cc) in cm and re.match(r"^\d{2}$", cm[(r0 - dr, cc)]):
                row_codes.append(cm[(r0 - dr, cc)])
                cc += 1
            if row_codes:
                colors = row_codes
                break
    if colors:
        table_rows.insert(0, ("", [int(x) for x in colors]))
    d["size_table"] = table_rows
    return [d]

def parse_solid_sheet(ws):
    """One sheet -> many designs (one per size x color combo with qty > 0)."""
    cm = cell_map(ws)
    base = parse_common(cm)
    # colors from legend like '09-BLACK'
    colors = {}
    for _, v in cm.items():
        m = re.match(r"^(\d{2})[-:]\s*([A-Z ]+)$", v)
        if m:
            colors[m.group(1)] = m.group(2).strip()
    # sizes like 'XS-002' / 'XS:002'
    sizes = {}
    for _, v in cm.items():
        m = re.match(r"^([A-Z0-9]{1,4})[-:](\d{3})$", v)
        if m and m.group(1).upper() in SIZE_NAMES:
            sizes[m.group(1).upper()] = m.group(2)
    # qty grid: header row containing color codes, size rows below
    hdr = None
    for (r, c), v in sorted(cm.items()):
        if re.match(r"^\d{2}:", v):  # e.g. 09:BLACK header cell
            hdr = r
            break
    grid = {}  # (size, color_code) -> qty
    if hdr:
        col_color = {}
        for (r, c), v in cm.items():
            if r == hdr:
                m = re.match(r"^(\d{2}):", v)
                if m:
                    col_color[c] = m.group(1)
        for (r, c), v in sorted(cm.items()):
            if r <= hdr:
                continue
            m = re.match(r"^([A-Z0-9]{1,4}):(\d{3})$", v)
            if m and m.group(1).upper() in SIZE_NAMES:
                size = m.group(1).upper()
                for cc, ccode in col_color.items():
                    q = cm.get((r, cc))
                    if q and re.match(r"^\d+$", q):
                        grid[(size, ccode)] = int(q)
    designs = []
    for (size, ccode), qty in sorted(grid.items()):
        d = dict(base)
        d["type"] = "solid"
        d["size"] = size
        d["color_code"] = ccode
        d["color_name"] = colors.get(ccode, "")
        d["code"] = f"{ccode}-{sizes.get(size,'???')}-000"
        d["qty_boxes"] = qty
        d["size_table"] = []
        designs.append(d)
    # order: colour first (the leading "09"/"69"/"78"), then size within it,
    # so pages print 09-003, 09-004, 09-005, 09-006, then 69-..., then 78-...
    def _key(d):
        parts = (d["code"] or "").split("-")
        colour = parts[0] if len(parts) > 0 else ""
        size = parts[1] if len(parts) > 1 else ""
        return (colour, size)
    designs.sort(key=_key)
    return designs

# ----- circle letter (C / D / B) auto-detection from embedded images -----
def _shape_letter(png_bytes):
    """Detect B/C/D from the ink SHAPE — no external OCR program needed.
       C is open on the right; D is closed on the right; B adds a middle bar."""
    try:
        from PIL import Image
        import numpy as np, io
        im = Image.open(io.BytesIO(png_bytes)).convert("L")
        a0 = np.array(im) < 100
        h, w = a0.shape
        # isolate the glyph inside the surrounding circle/oval ring
        inner = a0[int(h*0.25):int(h*0.75), int(w*0.25):int(w*0.75)]
        inner_im = Image.fromarray((~inner * 255).astype('uint8')).resize((60, 60))
        a = np.array(inner_im) < 128
        H, W = a.shape
        right = a[int(H*0.30):int(H*0.70), int(W*0.75):]
        left  = a[int(H*0.30):int(H*0.70), :int(W*0.25)]
        mid   = a[int(H*0.42):int(H*0.58), int(W*0.25):int(W*0.75)]
        is_open_right = right.mean() < 0.12
        if not is_open_right and mid.mean() > 0.3 and left.mean() > 0.2:
            return "B"
        return "C" if is_open_right else "D"
    except Exception:
        return None

def _ocr_letter_backup(png_bytes):
    """Optional OCR check if Tesseract happens to be installed. May return None."""
    try:
        import pytesseract
        from PIL import Image, ImageOps
        import io
        im = Image.open(io.BytesIO(png_bytes)).convert("L")
        w, h = im.size
        im = im.crop((int(w*0.22), int(h*0.22), int(w*0.78), int(h*0.78)))
        im = ImageOps.autocontrast(im).resize((120, 120))
        t = pytesseract.image_to_string(
            im, config="--psm 10 -c tessedit_char_whitelist=BCD").strip().upper()
        t = "".join(ch for ch in t if ch in "BCD")
        return t[0] if t else None
    except Exception:
        return None

def _ocr_letter(png_bytes):
    """Primary = shape detection (no install needed); OCR only breaks ties."""
    shape = _shape_letter(png_bytes)
    ocr = _ocr_letter_backup(png_bytes)
    if shape and ocr and shape == ocr:
        return shape
    return shape or ocr

def detect_sheet_letters(path):
    """Return {sheet_name: 'C'/'D'/'B'} by OCR-ing each sheet's circle image.
       The circle letter is a small square-ish PNG referenced by that sheet's
       drawing; we OCR every small image and keep the best letter per sheet."""
    result = {}
    try:
        z = zipfile.ZipFile(path)
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        # count how many drawings reference each media file
        from collections import Counter
        _img_usage = Counter()
        for dn in z.namelist():
            if re.match(r"xl/drawings/_rels/drawing\d+\.xml\.rels$", dn):
                for im in re.findall(r'Target="\.\./media/([^"]+)"',
                                     z.read(dn).decode()):
                    _img_usage[im] += 1
        # map worksheetN -> drawingN -> image targets
        for idx, sname in enumerate(names, start=1):
            rels = f"xl/worksheets/_rels/sheet{idx}.xml.rels"
            if rels not in z.namelist():
                continue
            drawing = None
            for m in re.finditer(r'Target="\.\./drawings/(drawing\d+\.xml)"',
                                 z.read(rels).decode()):
                drawing = m.group(1)
            if not drawing:
                continue
            drels = f"xl/drawings/_rels/{drawing}.rels"
            if drels not in z.namelist():
                continue
            imgs = re.findall(r'Target="\.\./media/([^"]+)"',
                              z.read(drels).decode())
            # The letter is the small PNG that DIFFERS between sheets.
            # Skip the big logo and the shared handling-symbol jpeg.
            best = None
            candidates = [im for im in imgs
                          if im.lower().endswith(".png")
                          and f"xl/media/{im}" in z.namelist()
                          and z.getinfo(f"xl/media/{im}").file_size <= 8000]
            # prefer a PNG unique to this sheet (not shared by others)
            from collections import Counter
            shared = {im for im, cnt in _img_usage.items() if cnt > 1}
            uniq = [im for im in candidates if im not in shared]
            ordered = uniq + [im for im in candidates if im in shared]
            for img in ordered:
                letter = _ocr_letter(z.read(f"xl/media/{img}"))
                if letter:
                    best = letter
                    break
            if best:
                result[sname] = best
    except Exception:
        pass
    return result

def detect_format(wb):
    for sn in wb.sheetnames:
        if "solid" in sn.lower():
            return "solid"
    return "ratio"

def parse_po(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    fmt = detect_format(wb)
    sheet_letters = detect_sheet_letters(path)   # {sheet_name: 'C'/'D'/'B'}
    designs = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        letter = sheet_letters.get(sn)
        try:
            if fmt == "solid":
                got = parse_solid_sheet(ws)
            else:
                got = parse_ratio_sheet(ws)
            for d in got:
                d["circle_letter"] = letter    # may be None -> fallback later
            designs += got
        except Exception as e:
            print(f"  WARNING: could not parse sheet '{sn}': {e}")
    return fmt, designs

# ----------------------------------------------------------------------
# PDF ASSEMBLY — page positions only; per-sticker drawing is in the
# layout_*.py files.
# ----------------------------------------------------------------------
def sheets_for_design(d):
    """How many printed sheets this design needs.
       cartons -> stickers (x STICKERS_PER_CARTON) -> sheets (/ copies-per-sheet).
       Returns (copies_per_sheet, stickers_needed, sheets_needed)."""
    n = copies_for_height(d["H"] or 999)
    cartons = d.get("qty_boxes") or 0
    stickers = cartons * STICKERS_PER_CARTON
    sheets = math.ceil(stickers / n) if stickers > 0 else 0
    return n, stickers, sheets

def _draw_one_page(c, d, n, show_dest):
    pw, ph = PAGE_SIZE
    lay = LAYOUTS.get(n, layout_one)
    cols, rows = lay.GRID
    sw = lay.STICKER_W_IN * 72
    sh = lay.STICKER_H_IN * 72
    grid_w = cols * sw
    y_top = ph                      # flush at the very top of the sheet
    if lay.PLACEMENT == "center":
        x0 = (pw - grid_w) / 2
    else:
        x0 = pw - grid_w - 2 * mm
    i = 0
    for r in range(rows):
        for col in range(cols):
            x = x0 + col * sw
            y = y_top - (r + 1) * sh
            lay.draw(c, x, y, sw, sh, d, show_dest=show_dest)
            i += 1
            if i >= n:
                break
    lay.page_lines(c, pw, ph, x0, y_top, sw, sh)
    c.showPage()

def generate_pdf(designs, out_path, po_label, show_dest=False,
                 one_page_each=False, circle_letter=None):
    """Generate the sticker PDF.
       one_page_each=False (default): repeat each design's page as many times
         as needed to cover its carton quantity; skip zero-qty designs.
       one_page_each=True: one master page per design (old behaviour)."""
    default_letter = circle_letter or common.CIRCLE_LETTER
    c = canvas.Canvas(out_path, pagesize=PAGE_SIZE)
    for d in designs:
        # per-sheet auto-detected letter wins; else the app/CLI default
        common.CIRCLE_LETTER = d.get("circle_letter") or default_letter
        n, stickers, sheets = sheets_for_design(d)
        if one_page_each:
            if (d.get("qty_boxes") or 0) <= 0:
                continue
            _draw_one_page(c, d, n, show_dest)
        else:
            for _ in range(sheets):        # 0 sheets -> skipped automatically
                _draw_one_page(c, d, n, show_dest)
    c.save()

def write_summary(all_rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    hdrs = ["File", "Type", "Code", "Item No", "Description", "PCS",
            "L", "W", "H", "Volume", "Copies/Sheet", "Qty (boxes)",
            "Stickers", "Sheets", "Size/Color"]
    ws.append(hdrs)
    from openpyxl.styles import Font as XFont
    for cell in ws[1]:
        cell.font = XFont(name="Arial", bold=True)
    for r in all_rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = XFont(name="Arial")
    widths = [26, 8, 12, 20, 34, 6, 6, 6, 6, 10, 12, 11, 9, 8, 16]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
    wb.save(out_path)

def main(paths, outdir="."):
    summary = []
    for path in paths:
        name = os.path.splitext(os.path.basename(path))[0]
        fmt, designs = parse_po(path)
        print(f"{name}: format={fmt}, {len(designs)} sticker designs")
        po_label = next((d["po_no"] for d in designs if d.get("po_no")), name)
        out_pdf = os.path.join(outdir, f"{name}_STICKERS.pdf")
        generate_pdf(designs, out_pdf, po_label)
        # sheet-count summary
        tot_cartons = tot_stickers = tot_sheets = 0
        print(f"  {'design':16}{'cartons':>8}{'stickers':>10}{'sheets':>8}")
        for d in designs:
            n, stickers, sheets = sheets_for_design(d)
            cartons = d.get('qty_boxes') or 0
            if cartons > 0:
                print(f"  {str(d['code']):16}{cartons:>8}{stickers:>10}{sheets:>8}")
            tot_cartons += cartons; tot_stickers += stickers; tot_sheets += sheets
        print(f"  {'TOTAL':16}{tot_cartons:>8}{tot_stickers:>10}{tot_sheets:>8}")
        print(f"  -> {out_pdf}  ({tot_sheets} pages)")
        for d in designs:
            n = copies_for_height(d["H"] or 999)
            sc = (f"{d.get('size','')} / {d.get('color_code','')}-{d.get('color_name','')}"
                  if d["type"] == "solid" else
                  ", ".join(f"{s}:{'/'.join(map(str,nums))}" for s, nums in d["size_table"] if s != "TOTAL"))
            _, stickers, sheets = sheets_for_design(d)
            summary.append([os.path.basename(path), d["type"], d["code"], d["item_no"],
                            d["description"], d["pcs"], d["L"], d["W"], d["H"],
                            d["volume"], n, d.get("qty_boxes"), stickers, sheets, sc])
    out_x = os.path.join(outdir, "STICKER_SUMMARY.xlsx")
    write_summary(summary, out_x)
    print(f"  -> {out_x}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1:])