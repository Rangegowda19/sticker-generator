"""
ratio_parser.py — reads a RATIO-type PO Excel and extracts, per sheet:
  dest, item_no, code (E-code = sheet name), pcs, description, volume,
  measurement (for height->layout), and the SIZE TABLE (list of (size, qty)).

Ratio stickers show a size breakdown table. The table orientation is decided
by how many sizes there are:
  - 1 size  -> single-cell style
  - many    -> full grid
Sticker box sizes match the SOLID type (8-up and 2-up only for ratio).
"""
import re
import openpyxl

# Cell map (from analysis of the ratio Excel), 1-indexed (row, col).
# col letters: AA=27, AB=28, AE=31, AF=32, P=16
CELL = {
    "dest":  (6, 32),    # AF6
    "item":  (12, 31),   # AE12
    "ecode": (13, 31),   # AE13  (also equals the sheet name)
    "pcs":   (14, 31),   # AE14
    "desc":  (15, 27),   # AA15
    "vol":   (23, 16),   # P23
    "meas":  (40, 7),    # G40 (measurement L..W..H..)
}
SIZE_LABEL_COL = 27   # AA
SIZE_QTY_COL   = 28   # AB
SIZE_ROW_RANGE = range(19, 30)

_SIZE_ORDER = ["XS","S","M","L","XL","XXL","3XL","XXXL","4XL"]

def _clean(v):
    return "" if v is None else str(v).strip()

def _num(s):
    m = re.search(r"\d+", str(s or ""))
    return int(m.group()) if m else 0

def parse_ratio_po(path):
    """Return list of design dicts, one per sheet (skipping empty/non-sticker)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    designs = []
    for name in wb.sheetnames:
        ws = wb[name]
        dest = _clean(ws.cell(*CELL["dest"]).value)
        item = _clean(ws.cell(*CELL["item"]).value)
        if not item and not dest:
            continue   # not a sticker sheet
        ecode = _clean(ws.cell(*CELL["ecode"]).value) or name
        pcs   = _num(ws.cell(*CELL["pcs"]).value)
        desc  = _clean(ws.cell(*CELL["desc"]).value)
        vol   = _clean(ws.cell(*CELL["vol"]).value)
        meas  = _clean(ws.cell(*CELL["meas"]).value)

        # size table
        sizes = []
        for r in SIZE_ROW_RANGE:
            lbl = _clean(ws.cell(r, SIZE_LABEL_COL).value)
            qty = ws.cell(r, SIZE_QTY_COL).value
            if lbl and lbl.lower() != "total":
                sizes.append((lbl, _num(qty)))
        total = sum(q for _, q in sizes)

        # measurement height -> copies-per-sheet (same rule as solid)
        H = None
        m = re.search(r"H(\d+)", meas or "")
        if m: H = int(m.group(1))

        designs.append({
            "sheet": name,
            "dest": dest,
            "item_no": item,
            "code": ecode,          # the E-code (I017, X053, etc.)
            "pcs": pcs,
            "description": desc,
            "volume": vol,
            "measurement": meas,
            "H": H,
            "sizes": sizes,         # list of (label, qty)
            "total": total,
            "table_type": "single" if len(sizes) <= 1 else "grid",
        })
    return designs

if __name__ == "__main__":
    import sys
    ds = parse_ratio_po(sys.argv[1] if len(sys.argv)>1
                        else "/mnt/user-data/uploads/P0727-489933-001_-_DF74_-_Ratio.xlsx")
    print(f"parsed {len(ds)} ratio stickers")
    for d in ds[:4]:
        print(f"\n  sheet={d['sheet']}  dest={d['dest']}  item={d['item_no']}")
        print(f"    code(E)={d['code']}  pcs={d['pcs']}  H={d['H']}  vol={d['volume']}")
        print(f"    desc={d['description']}")
        print(f"    table={d['table_type']}  sizes={d['sizes']}  total={d['total']}")